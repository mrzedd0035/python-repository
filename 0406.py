# -*- coding: utf-8 -*-
"""
Created on Thu Apr  8 19:20:55 2021

@author: User
"""

import openpyxl
import glob

atext = glob.glob('./data/*')
btext = [[] for _ in range(10)]

for x in atext:
	xfd = openpyxl.load_Workbook(x, data_only=True)
	sheet = xfd['Sheet1']

	studentDict = {
		'st_num' : sheet['a3'].v,
		'name' : sheet['b3'].v,
		'group_id' : sheet['c3'].v,
		'git' : sheet['d3'].v
	}
	group = int(studentDict['group_id'][1:])
	btext[group-1].append(studentDict)
file = openpyxl.Workbook()
file.remove_sheet(file.get_sheet_by_name('Sheet'))

for i in range(10):
	newSheet = file.create_sheet()
	newSheet.title = 'Sheet%d'%(i+1)
	newSheet.append(['st_num', 'name', 'group_id', 'git'])

	for s in btext[i]:
		newSheet.append([str(s['st_num']), s['name'], s['group_id'], s['git']])

file.save('./Python_example.xlsx')